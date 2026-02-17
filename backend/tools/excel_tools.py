"""Ferramentas de Excel para agentes."""

from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path

from crewai_tools import tool
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Border, Font, PatternFill, Side

OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "outputs"))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


@tool("create_excel_spreadsheet")
def create_excel_spreadsheet(data_json: str, title: str = "Relatório", filename: str = "") -> str:
    """Cria planilha profissional com estilos e gráfico quando possível."""
    try:
        data = json.loads(data_json)
        if not isinstance(data, list) or not data:
            return "Erro: data_json deve ser uma lista de objetos com dados."
            return json.dumps({"success": False, "error": "data_json deve ser uma lista de objetos com dados."}, ensure_ascii=False)

        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"

        headers = list(data[0].keys())
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1F4E79")
        alt_fill = PatternFill("solid", fgColor="D9E1F2")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for row_idx, item in enumerate(data, start=2):
            ws.append([item.get(h) for h in headers])
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 50)

        numeric_cols = [idx + 1 for idx, h in enumerate(headers) if any(isinstance(x.get(h), (int, float)) for x in data)]
        if numeric_cols:
            chart_sheet = wb.create_sheet("Gráficos")
            chart = BarChart()
            chart.title = title
            chart.y_axis.title = "Valor"
            chart.x_axis.title = headers[0]
            values = Reference(ws, min_col=numeric_cols[0], min_row=1, max_row=len(data) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(data) + 1)
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(cats)
            chart_sheet.add_chart(chart, "A1")

        safe_filename = filename or f"relatorio_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        if not safe_filename.endswith(".xlsx"):
            safe_filename += ".xlsx"

        output_path = OUTPUT_DIR / safe_filename
        wb.save(output_path)
        return str(output_path)
    except Exception as exc:
        return f"Erro ao criar planilha: {exc}"
        return json.dumps(
            {
                "success": True,
                "message": "Planilha criada com sucesso.",
                "artifact_id": output_path.name,
                "artifact_path": str(output_path),
                "format": "xlsx",
            },
            ensure_ascii=False,
        )
    except Exception as exc:
        return json.dumps({"success": False, "error": f"Erro ao criar planilha: {exc}"}, ensure_ascii=False)
